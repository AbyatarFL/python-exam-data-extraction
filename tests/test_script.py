import sys
import os
import json
import sys
import openpyxl
import csv
import pytest

# add parent folder (Task-2) to sys.path
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from python_exam_data_extraction import (
    extract_address_components,
    extract_property_details,
    save_to_csv,
    save_to_json,
    save_to_xlsx,
    ensure_output_folder
)

def test_extract_address_components():
    text = "Location: 742 Evergreen Terrace, Springfield, ST 12345, United States"
    result = extract_address_components(text)
    assert result["street_number"] == "742"
    assert result["street_name"] == "Evergreen Terrace"
    assert result["city"] == "Springfield"
    assert result["state"] == "ST"
    assert result["zip_code"] == "12345"
    assert result["country"] == "United States"


def test_extract_property_details():
    text = """
    PROPERTY DETAILS - ID: PRO-2024-0123 
    Location: 742 Evergreen Terrace, Springfield, ST 12345, United States 
    Contact: Jane Doe (Primary) | john.doe@email.com 
    Additional Details: Built 1983 | Sq Ft: 2,400 | Lot: 0.25 acres 
    """
    result = extract_property_details(text)
    assert result["property_id"] == "PRO-2024-0123"
    assert result["contact_name"] == "Jane Doe"
    assert result["contact_type"] == "Primary"
    assert result["contact_email"] == "john.doe@email.com"
    assert result["year_built"] == "1983"
    assert result["square_footage"] == "2400"
    assert result["lot_size"] == "0.25 acres"
    assert result["city"] == "Springfield"


def test_save_to_csv(tmp_path):
    data = {"kolom1": 1, "kolom2": 2}
    file_path = tmp_path / "unit_test.csv"
    save_to_csv(data, file_path.name)

    # Open and check contents
    with open(os.path.join("Output_Files", file_path.name), newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        assert rows[0]["kolom1"] == "1"
        assert rows[0]["kolom2"] == "2"


def test_save_to_json(tmp_path):
    data = {"kolom_baru_1": 10, "kolom_baru_2": 20}
    file_path = tmp_path / "unit_test.json"
    save_to_json(data, file_path.name)

    with open(os.path.join("Output_Files", file_path.name), encoding="utf-8") as f:
        loaded = json.load(f)
        assert loaded["kolom_baru_1"] == 10
        assert loaded["kolom_baru_2"] == 20


def test_save_to_xlsx(tmp_path):
    data = {"test": "low", "angka": 123}
    file_path = tmp_path / "unit_test.xlsx"
    save_to_xlsx(data, file_path.name)

    wb = openpyxl.load_workbook(os.path.join("Output_Files", file_path.name))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    assert headers == list(data.keys())
    assert values == list(data.values())


def test_ensure_output_folder():
    folder = ensure_output_folder()
    assert os.path.exists(folder)